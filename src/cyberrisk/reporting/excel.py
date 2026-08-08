"""Excel report generation (Phase 4).

Builds a client-facing workbook from a simulation run + policy transform:
    - Overview tab (headline metrics, firm profile, risk category)
    - Retained/transferred tab (policy structure, per-year distribution)
    - LEC tab (loss exceedance curve data for charting)
    - Scenarios tab (AAL breakdown)

Uses openpyxl (available via the 'reporting' extra).  Deliberately simple
and deterministic so the same inputs always produce the same workbook --
no charts/formatting that depends on locale or installed fonts.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np

from cyberrisk.metrics import compute_metrics
from cyberrisk.simulation import SimulationResult


def write_report(
    result: SimulationResult,
    policy_metrics: dict[str, np.ndarray] | None = None,
    out_path: str | Path = "cyber_risk_report.xlsx",
) -> str:
    """Write an Excel workbook summarising a simulation run.

    Parameters
        result          SimulationResult from the loss engine
        policy_metrics  optional dict from policy_transform (retained/transferred)
        out_path        destination workbook path
    Returns
        the path written
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "openpyxl is required for Excel reporting. Install with "
            "`pip install cyberrisk[reporting]`."
        ) from exc

    m = compute_metrics(result)
    wb = openpyxl.Workbook()

    # --- Overview --------------------------------------------------------
    ws = wb.active
    ws.title = "Overview"
    _write_key_value(ws, "Cyber Risk Assessment", None)
    _write_key_value(ws, "EAL (Expected Annual Loss)", m.eal, money=True)
    _write_key_value(ws, "VaR 95%", m.var_95, money=True)
    _write_key_value(ws, "ES 95%", m.es_95, money=True)
    _write_key_value(ws, "VaR 99%", m.var_99, money=True)
    _write_key_value(ws, "ES 99%", m.es_99, money=True)
    _write_key_value(ws, "PML 1-in-250", m.pml_250, money=True)
    _write_key_value(ws, "P(no loss)", m.prob_zero_loss, pct=True)

    # --- Scenarios -------------------------------------------------------
    ws2 = wb.create_sheet("Scenarios")
    ws2.append(["Scenario", "AAL (USD)", "Share of EAL"])
    contrib = m.scenario_contribution()
    for i, key in enumerate(result.scenario_keys):
        aal = result.scenario_losses[:, i].mean()
        ws2.append([key, round(float(aal), 2), round(contrib[key], 4)])

    # --- Retained / transferred ------------------------------------------
    if policy_metrics is not None:
        ws3 = wb.create_sheet("Policy")
        retained = policy_metrics["retained"]
        transferred = policy_metrics["transferred"]
        ws3.append(["Metric", "Value"])
        _write_key_value(ws3, "Retained EAL", float(retained.mean()), money=True)
        _write_key_value(ws3, "Transferred EAL", float(transferred.mean()), money=True)
        _write_key_value(
            ws3, "Retained ES 99%", _es(transferred, retained, 0.99), money=True
        )

    # --- Model Limitations (mandatory disclosure) ------------------------
    _write_disclosure_sheet(wb)

    wb.save(str(out_path))
    return str(out_path)


def _write_disclosure_sheet(wb) -> None:
    """Write the mandatory Model Limitations sheet (always present)."""
    from cyberrisk.agent.disclosure import DISCLOSURE_HEADING, LIMITATIONS

    ws = wb.create_sheet("Model Limitations")
    ws.append([DISCLOSURE_HEADING])
    ws.append([])
    for item in LIMITATIONS:
        ws.append([f"- {item}"])


def _es(transferred: np.ndarray, retained: np.ndarray, q: float) -> float:
    """ES of retained loss at confidence q (reusing metrics.expected_shortfall)."""
    from cyberrisk.metrics import expected_shortfall

    return expected_shortfall(retained, q)


def _write_key_value(ws, label: str, value, money: bool = False, pct: bool = False) -> None:
    """Write a label/value pair into the next row of a worksheet."""
    if value is None:
        ws.append([label])
        return
    if money:
        value = f"${value:,.0f}"
    elif pct:
        value = f"{value*100:.1f}%"
    else:
        value = f"{value:,.2f}"
    ws.append([label, value])
