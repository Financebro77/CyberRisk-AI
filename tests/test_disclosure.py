"""Tests for the mandatory Model Limitations disclosure.

Every final advisory report -- whether produced by the DeepSeek consultant
agent, the rule-based / LLM consultant, or the Excel workbook -- must end with
the Model Limitations block containing all five mandated bullets.  These tests
lock that in and verify the block is never duplicated.
"""

from __future__ import annotations

import numpy as np

from cyberrisk.agent.disclosure import (
    DISCLOSURE_HEADING,
    LIMITATIONS,
    append_disclosure,
    disclosure_block,
)

# The five mandated statements, in order.
EXPECTED_LIMITATIONS = (
    "Cyber losses are probabilistic estimates, not predictions.",
    "Results depend on benchmark datasets and modelling assumptions.",
    "Catastrophic systemic cyber events may not be fully captured.",
    "Parameter uncertainty exists.",
    "Insurance terms and policy wording may affect actual recovery.",
)


# ---------------------------------------------------------------------------
# disclosure_block
# ---------------------------------------------------------------------------


def test_disclosure_block_has_heading_and_five_bullets():
    block = disclosure_block()
    assert block.startswith(DISCLOSURE_HEADING)
    assert len(LIMITATIONS) == 5
    assert LIMITATIONS == EXPECTED_LIMITATIONS
    for item in LIMITATIONS:
        assert f"- {item}" in block


def test_disclosure_matches_mandated_wording():
    """The exact bullets from the spec appear verbatim."""
    assert EXPECTED_LIMITATIONS == LIMITATIONS


def test_append_disclosure_is_idempotent():
    report = "Your advisory report text."
    once = append_disclosure(report)
    twice = append_disclosure(once)
    assert once == twice
    assert once.endswith(disclosure_block())
    assert once.count("Model Limitations") == 1


def test_append_disclosure_empty_report():
    assert append_disclosure("") == disclosure_block()


# ---------------------------------------------------------------------------
# Every final-report surface
# ---------------------------------------------------------------------------


def test_controller_final_answer_ends_with_disclosure():
    from cyberrisk.agent.agent_controller import CyberRiskAgent
    from cyberrisk.agent.deepseek_client import ChatResponse
    from cyberrisk.agent.memory import ConversationMemory
    from cyberrisk.agent.schemas import AgentConfig

    class ScriptedClient:
        def __init__(self):
            self.script = [ChatResponse(content="Here is your assessment.")]

        def chat(self, messages, tools=None, temperature=None, max_tokens=None):
            return self.script.pop(0)

    agent = CyberRiskAgent(
        client=ScriptedClient(), config=AgentConfig(), memory=ConversationMemory()
    )
    answer = agent.chat("Assess us", welcome=True)
    assert answer.endswith(disclosure_block())
    assert "probabilistic estimates, not predictions" in answer
    # Memory stores the disclosed answer, so follow-up turns do not duplicate it.
    assert agent.memory.get()[-1]["content"].endswith(disclosure_block())


def test_rule_based_recommendation_carries_disclosure():
    from agent.consultant_agent import generate_recommendations
    from cyberrisk.calibration import load_config
    from cyberrisk.metrics import compute_metrics
    from cyberrisk.scoring import CompanyProfile, compute_score
    from cyberrisk.simulation import simulate
    from pathlib import Path

    repo = Path(__file__).parent.parent
    cfg = load_config(repo / "config" / "scenarios.yaml", repo / "config" / "simulation_config.yaml")
    profile = CompanyProfile(firm_name="Acme Corp", factor_scores={
        "external_attack_surface": 90.0, "industry_targeting": 85.0,
        "data_sensitivity": 80.0, "patch_cadence": 90.0, "mfa_coverage": 90.0,
        "edr_coverage": 85.0, "backup_frequency": 60.0, "vendor_assessment": 70.0,
    })
    scored = compute_score(profile)
    result = simulate(cfg, n_years=10_000, score=scored.composite_score)
    m = compute_metrics(result)
    rec = generate_recommendations(scored, m)
    assert rec.disclosure == disclosure_block()
    report = rec.full_report()
    assert report.endswith(disclosure_block())
    assert report.count("Model Limitations") == 1


def test_excel_report_has_disclosure_sheet(tmp_path):
    from cyberrisk.calibration import load_config
    from cyberrisk.reporting.excel import write_report
    from cyberrisk.simulation import simulate
    from pathlib import Path

    repo = Path(__file__).parent.parent
    cfg = load_config(repo / "config" / "scenarios.yaml", repo / "config" / "simulation_config.yaml")
    result = simulate(cfg, n_years=5_000)
    out = tmp_path / "report.xlsx"
    write_report(result, out_path=out)

    import openpyxl

    wb = openpyxl.load_workbook(out)
    assert "Model Limitations" in wb.sheetnames
    ws = wb["Model Limitations"]
    rows = [r[0].value for r in ws.iter_rows() if r[0].value]
    assert rows[0] == "Model Limitations"
    for item in EXPECTED_LIMITATIONS:
        assert f"- {item}" in rows, item


# ---------------------------------------------------------------------------
# Prompts enforce the disclosure
# ---------------------------------------------------------------------------


def test_system_prompt_mandates_disclosure():
    from cyberrisk.agent.prompts import GROUNDING_REMINDER, SYSTEM_PROMPT

    assert "MANDATORY MODEL-LIMITATIONS DISCLOSURE" in SYSTEM_PROMPT
    assert "probabilistic estimates, not predictions" in SYSTEM_PROMPT
    assert "Model Limitations" in GROUNDING_REMINDER
    assert "probabilistic estimates, not predictions" in GROUNDING_REMINDER


def test_rule_based_prompts_mandate_disclosure():
    from agent.prompts import SAFETY_SYSTEM_PROMPT, SYSTEM_PROMPT

    assert "Model Limitations" in SYSTEM_PROMPT
    assert "probabilistic estimates, not predictions" in SYSTEM_PROMPT
    assert "Model Limitations" in SAFETY_SYSTEM_PROMPT
    assert "probabilistic estimates, not predictions" in SAFETY_SYSTEM_PROMPT
